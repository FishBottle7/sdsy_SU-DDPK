import openpyxl
import os
import atexit

# Load the workbook
def load_workbook():
    path = input("输入表格路径：")
    path = path.strip('\'"')
    while True:
        try:
            wb = openpyxl.load_workbook(path)
            break
        except PermissionError:
            input("请关闭文件后重试")
    return wb, path
class WorkbookProcessor:
    def __init__(self,wb,path):
        self.wb = wb
        self.path = path
        self.sheet = self.wb.active
        self.save_path = None
        atexit.register(self.exit_handler)



    #拆分合并单元格
    def unmerge_cells(self):
        merged_cells_ranges = list(self.wb.active.merged_cells.ranges)  # Create a copy

        # Loop through all merged cells
        for merged_cells_range in merged_cells_ranges:
            # Get the top-left cell value
            top_left_cell_value = merged_cells_range.start_cell.value

            # Unmerge cells
            self.wb.active.unmerge_cells(str(merged_cells_range))

            # Fill unmerged cells with the top-left cell value
            for row in self.wb.active[merged_cells_range.coord]:
                for cell in row:
                    cell.value = top_left_cell_value


    def save_workbook(self):
        file_path, extension = os.path.splitext(self.path)

        # Add 'processed' to the file path
        new_file_path = f"{file_path}_processed{extension}"

        # Save the workbook with the new file path
        try:
            self.wb.save(new_file_path)
        except PermissionError:
            input("请关闭文件后重试")
            return self.save_workbook(self.wb, self.path)
        self.save_path = new_file_path
        print(f"Workbook saved at {new_file_path}")

    def keep_max_rows(self):
        max_rows = {}
        max_seq = {}  # Initialize max sequence number for each name
        for row in self.sheet.iter_rows(min_row=2, values_only=True):  # Skip header row
            name = row[1]
            try:
                seq = int(row[0])  # Convert '序号' to int
            except Exception:
                continue  # Skip rows where '序号' cannot be converted to int
            if name not in max_seq or seq > max_seq[name]:  # If sequence number is greater than max sequence number for this name
                max_rows[name] = [row]  # Reset rows for this name
                max_seq[name] = seq  # Update max sequence number for this name
            elif seq == max_seq[name]:  # If sequence number is equal to max sequence number for this name
                max_rows[name].append(row)  # Add row to list of rows for this name

        # Create a new workbook with only the max rows
        new_wb = openpyxl.Workbook()
        new_sheet = new_wb.active
        new_sheet.append(list(next(self.sheet.iter_rows(min_row=1, max_row=1, values_only=True))))  # Copy header row
        for rows in max_rows.values():
            for row in rows:
                new_sheet.append(row)
        self.wb = new_wb
        self.sheet = self.wb.active

    def remove_zeros(self):
        rows_to_delete = []  # List to keep track of rows to delete

        # Iterate over rows
        for i, row in enumerate(self.sheet.iter_rows(min_row=2, values_only=True), start=2):  # Skip header row
            if float(row[4]) == 0:  # If value in the fifth column is 0
                rows_to_delete.append(i)

        # Delete rows in reverse order to avoid shifting indices
        for i in reversed(rows_to_delete):
            self.sheet.delete_rows(i)

    def absolute_values(self):

        for row in self.sheet.iter_rows(min_row=2):  # Skip header row
            try:
                value = float(row[4].value)
                row[4].value = abs(value)
            except ValueError:
                continue


    def calculate_scores(self):

        # Set headers for columns 13 and 14
        self.sheet.cell(row=1, column=13, value="班级")
        self.sheet.cell(row=1, column=14, value="总分")

        # Initialize class names and scores
        class_names = [f"高一（{i}）班" for i in range(1, 14)] + [f"高二（{i}）班" for i in range(1, 14)] + [f"高三（{i}）班" for i in range(1, 14)]
        class_scores = {class_name: 0 for class_name in class_names}

        # Calculate scores for each class
        for row in self.sheet.iter_rows(min_row=2, values_only=True):  # Skip header row
            grade = row[2]
            class_num = row[3]
            score = float(row[4])
            class_name = f"{grade}（{class_num}）班"
            if class_name in class_scores:
                class_scores[class_name] -= score  # Subtract score because we want the opposite of the sum

        # Write class names and scores to columns 13 and 14
        for i, (class_name, score) in enumerate(class_scores.items(), start=2):  # Skip header row
            self.sheet.cell(row=i, column=13, value=class_name)
            self.sheet.cell(row=i, column=14, value=score)


    def calculate_scores_with_classify(self):


        # Set headers for columns 13, 14, 15, 16, 17
        self.sheet.cell(row=1, column=13, value="班级")
        self.sheet.cell(row=1, column=14, value="升旗")
        self.sheet.cell(row=1, column=15, value="两操")
        self.sheet.cell(row=1, column=16, value="日常")
        self.sheet.cell(row=1, column=17, value="周五检查")

        # Initialize class names and scores
        class_names = [f"高一（{i}）班" for i in range(1, 14)] + ["1"] + [f"高二（{i}）班" for i in range(1, 14)] + ["2"] + [f"高三（{i}）班"for i in range(1, 14)]
        class_category_scores = {class_name: {"升旗": None, "两操": None, "日常": None, "周五检查": None} if class_name == "1" or class_name == "2"
                                            else {"升旗": 0, "两操": 0, "日常": 0, "周五检查": 0} for class_name in class_names}
        # Calculate scores for each class
        for row in self.sheet.iter_rows(min_row=2, values_only=True):  # Skip header row
            grade = row[2]
            class_num = row[3].replace("班", "")  # Remove '班' from class number
            score = float(row[4])
            category = row[6]
            class_name = f"{grade}（{class_num}）班"
            if class_name in class_category_scores:
                if category in class_category_scores[class_name]:
                    class_category_scores[class_name][category] -= score  # Subtract score because we want the opposite of the sum
                else:
                    print(f"line {row} Unexpected category: {category}")
                    continue
            else:
                print(f"line {row} Unexpected class name: {class_name}")
                continue

        # Write class names to column 13
        for i, class_name in enumerate(class_names, start=2):  # Skip header row
            self.sheet.cell(row=i, column=13, value=class_name)


        # Write class category scores to columns 14, 15, 16, 17
        for i, (class_name, category_scores) in enumerate(class_category_scores.items(), start=2):  # Skip header row
            self.sheet.cell(row=i, column=14, value=category_scores["升旗"])
            self.sheet.cell(row=i, column=15, value=category_scores["两操"])
            self.sheet.cell(row=i, column=16, value=category_scores["日常"])
            self.sheet.cell(row=i, column=17, value=category_scores["周五检查"])


    def format_grade_and_class(self):
        grade_mapping = {"高一": "1", "高二": "2", "高三": "3"}  # Add a mapping from grade names to numbers

        # Iterate over rows
        for row in self.sheet.iter_rows(min_row=2):  # Skip header row
            grade = grade_mapping.get(row[2].value, row[2].value)  # Use the mapping to convert grade names to numbers
            #class_num = row[3].value.replace("班", "")  # Remove '班' from class number
            class_num = row[3].value.replace("班", "") if row[3].value else None
            if class_num is None: continue
            formatted_value = f"{grade},{class_num}"
            row[7].value = formatted_value  # Write to the 8th column (0-indexed)


    def exit_handler(self):
        input("按任意键退出")

    def run(self):
        self.unmerge_cells()
        self.keep_max_rows()
        self.remove_zeros()
        self.absolute_values()
        self.calculate_scores_with_classify()
        self.format_grade_and_class()
        self.save_workbook()

if __name__ == "__main__":
    processor = WorkbookProcessor(*load_workbook())
    processor.run()
    os.startfile(processor.save_path)